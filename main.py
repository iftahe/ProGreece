from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, text
from typing import List, Optional
from datetime import datetime
from decimal import Decimal
from fastapi.middleware.cors import CORSMiddleware
import json
import os
import models, schemas
import services.budget_report_service
import services.forecast_service
from services.export_service import export_to_excel
from services.audit_service import log_audit
from database import SessionLocal, engine, DB_NAME, IS_RENDER


def _clean_date(val: Optional[str]) -> Optional[str]:
    """Normalize empty/whitespace date strings to None."""
    return val if val and val.strip() else None


# Create tables (only if they don't exist)
models.Base.metadata.create_all(bind=engine)

# Run phase-4 migration (idempotent) to add new columns to existing tables
# create_all only creates NEW tables; it won't ALTER existing ones.
from migrations.phase4_migrate import main as run_phase4_migration
try:
    run_phase4_migration()
except Exception as exc:
    print(f"[startup] phase4 migration note: {exc}")

from migrations.plan_versioning_migrate import main as run_plan_versioning_migration
try:
    run_plan_versioning_migration()
except Exception as exc:
    print(f"[startup] plan versioning migration note: {exc}")

# Create performance indexes
with engine.connect() as conn:
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_tx_project_date ON transactions(project_id, date)"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_tx_status_date ON transactions(status, date)"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_tx_customer ON transactions(customer_id_fk)"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_tx_invoice ON transactions(invoice_id)"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_tx_counterparty ON transactions(counterparty_id)"))
    conn.commit()

app = FastAPI()

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# --- Routes ---

@app.get("/")
def read_root():
    return {"message": "ProGreece API is running"}

@app.get("/health")
def health_check():
    """Health check endpoint for deployment diagnostics."""
    status = {"api": "ok", "render": bool(IS_RENDER), "db_path": DB_NAME}
    try:
        db_exists = os.path.exists(DB_NAME)
        status["db_file_exists"] = db_exists
        if db_exists:
            status["db_size_bytes"] = os.path.getsize(DB_NAME)
        db = SessionLocal()
        count = db.query(models.Project).count()
        status["db_connected"] = True
        status["project_count"] = count

        # Duplicate apartment audit
        dup_query = db.query(
            models.Apartment.project_id,
            models.Apartment.apartment_number,
            func.count(models.Apartment.id).label("cnt")
        ).filter(
            models.Apartment.apartment_number.isnot(None),
            models.Apartment.apartment_number != "",
        ).group_by(
            models.Apartment.project_id,
            models.Apartment.apartment_number,
        ).having(func.count(models.Apartment.id) > 1).all()

        extra_copies = sum(row.cnt - 1 for row in dup_query)
        status["duplicate_apartments"] = extra_copies
        status["duplicate_apartment_groups"] = [
            {"project_id": row.project_id, "apartment_number": row.apartment_number, "count": row.cnt}
            for row in dup_query
        ]

        db.close()
    except Exception as e:
        status["db_connected"] = False
        status["db_error"] = str(e)
    return status

@app.get("/projects/", response_model=List[schemas.Project])
def read_projects(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    projects = db.query(models.Project).offset(skip).limit(limit).all()
    
    # Calculate total_budget dynamically for each project
    result = []
    for project in projects:
        # Sum all planned_amount from budget_categories for this project
        total_budget = db.query(func.sum(models.BudgetCategory.planned_amount)).filter(
            models.BudgetCategory.project_id == project.id
        ).scalar() or 0
        
        # Create a dict with all project fields and override total_budget
        project_dict = {
            "id": project.id,
            "name": project.name,
            "status": project.status,
            "project_account_val": float(project.project_account_val) if project.project_account_val else 0,
            "property_cost": float(project.property_cost) if project.property_cost else None,
            "remarks": project.remarks,
            "account_balance": float(project.account_balance) if project.account_balance else 0,
            "total_budget": float(total_budget) if total_budget else None
        }
        result.append(project_dict)
    
    return result

@app.post("/projects/", response_model=schemas.Project)
def create_project(project: schemas.ProjectCreate, db: Session = Depends(get_db)):
    db_project = models.Project(
        name=project.name, 
        status=project.status or "Active",
        project_account_val=project.project_account_val or 0,
        property_cost=project.property_cost,
        remarks=project.remarks,
        account_balance=project.account_balance or 0,
        total_budget=project.total_budget
    )
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    return db_project

@app.put("/projects/{project_id}", response_model=schemas.Project)
def update_project(project_id: int, project: schemas.ProjectCreate, db: Session = Depends(get_db)):
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    db_project.name = project.name
    db_project.status = project.status or "Active"
    db_project.project_account_val = project.project_account_val or 0
    db_project.property_cost = project.property_cost
    db_project.remarks = project.remarks
    db_project.account_balance = project.account_balance or 0
    db_project.total_budget = project.total_budget
    
    db.commit()
    db.refresh(db_project)
    return db_project

@app.get("/transactions/")
def read_transactions(
    skip: int = 0,
    limit: int = 50,
    project_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    transaction_type: Optional[int] = None,
    tx_type: Optional[str] = None,
    budget_item_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.Transaction)
    if project_id:
        query = query.filter(models.Transaction.project_id == project_id)
    if budget_item_id:
        query = query.filter(models.Transaction.budget_item_id == budget_item_id)
    if date_from:
        query = query.filter(models.Transaction.date >= datetime.fromisoformat(date_from))
    if date_to:
        dt_to = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59)
        query = query.filter(models.Transaction.date <= dt_to)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                models.Transaction.remarks.ilike(pattern),
                models.Transaction.description.ilike(pattern),
                models.Transaction.supplier.ilike(pattern),
                models.Transaction.category.ilike(pattern),
            )
        )
    if transaction_type is not None:
        query = query.filter(models.Transaction.transaction_type == transaction_type)
    if tx_type:
        query = query.filter(models.Transaction.type == tx_type)
    total = query.count()
    transactions = query.order_by(models.Transaction.date.desc()).offset(skip).limit(limit).all()
    return {"items": transactions, "total": total, "skip": skip, "limit": limit}

@app.post("/transactions/", response_model=schemas.Transaction)
def create_transaction(transaction: schemas.TransactionCreate, db: Session = Depends(get_db)):
    # Handle VAT logic: if from_account or to_account is system account, set vat_rate to 0
    vat_rate = transaction.vat_rate or 0
    if transaction.from_account_id:
        from_acc = db.query(models.Account).filter(models.Account.id == transaction.from_account_id).first()
        if from_acc and from_acc.is_system_account:
            vat_rate = 0
    if transaction.to_account_id:
        to_acc = db.query(models.Account).filter(models.Account.id == transaction.to_account_id).first()
        if to_acc and to_acc.is_system_account:
            vat_rate = 0

    transaction_data = transaction.dict()
    transaction_data['vat_rate'] = vat_rate
    db_transaction = models.Transaction(**transaction_data)
    db.add(db_transaction)

    # Auto-compute VAT and withholding amounts
    _vat_rate = Decimal(str(vat_rate)) if vat_rate else Decimal('0')
    _withholding_rate = Decimal(str(transaction.withholding_rate)) if transaction.withholding_rate else Decimal('0')
    _amount = Decimal(str(transaction.amount)) if transaction.amount else Decimal('0')
    db_transaction.vat_amount = _amount * _vat_rate
    db_transaction.withholding_amount = _amount * _withholding_rate

    db.commit()
    db.refresh(db_transaction)

    # Audit log
    log_audit(db, "transaction", db_transaction.id, "create", None, transaction.dict())

    # Feature 3: Upsert AccountCategoryMapping
    _upsert_account_category_mapping(db, db_transaction)

    return db_transaction

@app.put("/transactions/{transaction_id}", response_model=schemas.Transaction)
def update_transaction(transaction_id: int, transaction: schemas.TransactionCreate, db: Session = Depends(get_db)):
    db_transaction = db.query(models.Transaction).filter(models.Transaction.id == transaction_id).first()
    if not db_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # Capture old data before update
    old_data = {c.name: str(getattr(db_transaction, c.name)) for c in models.Transaction.__table__.columns}

    # Handle VAT logic
    vat_rate = transaction.vat_rate or 0
    if transaction.from_account_id:
        from_acc = db.query(models.Account).filter(models.Account.id == transaction.from_account_id).first()
        if from_acc and from_acc.is_system_account:
            vat_rate = 0
    if transaction.to_account_id:
        to_acc = db.query(models.Account).filter(models.Account.id == transaction.to_account_id).first()
        if to_acc and to_acc.is_system_account:
            vat_rate = 0

    transaction_data = transaction.dict()
    transaction_data['vat_rate'] = vat_rate

    for key, value in transaction_data.items():
        setattr(db_transaction, key, value)

    # Auto-compute VAT and withholding amounts
    _vat_rate = Decimal(str(vat_rate)) if vat_rate else Decimal('0')
    _withholding_rate = Decimal(str(transaction.withholding_rate)) if transaction.withholding_rate else Decimal('0')
    _amount = Decimal(str(transaction.amount)) if transaction.amount else Decimal('0')
    db_transaction.vat_amount = _amount * _vat_rate
    db_transaction.withholding_amount = _amount * _withholding_rate

    db.commit()
    db.refresh(db_transaction)

    # Audit log
    log_audit(db, "transaction", transaction_id, "update", old_data, transaction.dict())

    # Feature 3: Upsert AccountCategoryMapping
    _upsert_account_category_mapping(db, db_transaction)

    return db_transaction

@app.delete("/transactions/{transaction_id}")
def delete_transaction(transaction_id: int, db: Session = Depends(get_db)):
    db_transaction = db.query(models.Transaction).filter(models.Transaction.id == transaction_id).first()
    if not db_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # Capture old data before delete
    old_data = {c.name: str(getattr(db_transaction, c.name)) for c in models.Transaction.__table__.columns}

    db.delete(db_transaction)
    db.commit()

    # Audit log (entity_id is still valid as a reference)
    log_audit(db, "transaction", transaction_id, "delete", old_data, None)

    return {"message": "Transaction deleted successfully"}

# --- דוחות ותקציב ---

@app.get("/reports/budget/{project_id}")
def get_budget_report(project_id: int):
    # שימוש בפונקציה החדשה והנכונה מה-Service
    try:
        return services.budget_report_service.get_budget_report(project_id)
    except Exception as e:
        print(f"Error generating budget report: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/reports/cash-flow/{project_id}")
def get_cash_flow_forecast(project_id: int, db: Session = Depends(get_db)):
    """תחזית תזרים מזומנים לפרויקט"""
    try:
        return services.forecast_service.generate_cash_flow_forecast(db, project_id)
    except Exception as e:
        print(f"Error generating cash flow forecast: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/reports/forecast/company")
def get_company_forecast(format: str = Query(None), db: Session = Depends(get_db)):
    """
    Company-wide 12-month consolidated cash flow forecast.

    Returns a 12-month forward window (from today) across all active projects,
    with monthly inflows, outflows, net, cumulative_cash and a cash_buffer_alert flag.
    """
    try:
        forecast_result = services.forecast_service.generate_company_forecast(db)
        if format == "xlsx":
            rows = forecast_result if isinstance(forecast_result, list) else forecast_result.get("rows", [])
            buf = export_to_excel("Company Forecast", rows, None, None)
            return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": "attachment; filename=Company Forecast.xlsx"})
        return forecast_result
    except Exception as e:
        print(f"Error generating company forecast: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/reports/forecast/projects")
def get_projects_forecast(format: str = Query(None), db: Session = Depends(get_db)):
    """
    Per-project 12-month cash flow forecast comparison.

    Returns a 12-month forward window (from today) broken down per project,
    including next_3_months_net, next_6_months_net, next_12_months_net and
    lowest_cash_point for each project.
    """
    try:
        forecast_result = services.forecast_service.generate_projects_forecast(db)
        if format == "xlsx":
            rows = forecast_result if isinstance(forecast_result, list) else forecast_result.get("rows", [])
            buf = export_to_excel("Projects Forecast", rows, None, None)
            return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": "attachment; filename=Projects Forecast.xlsx"})
        return forecast_result
    except Exception as e:
        print(f"Error generating projects forecast: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/reports/forecast/drilldown/{project_id}/{month}")
def get_forecast_drilldown(project_id: int, month: str, db: Session = Depends(get_db)):
    """Return individual items composing a forecast month.
    month format: YYYY-MM
    """
    from decimal import Decimal as D
    import calendar

    parts = month.split("-")
    if len(parts) != 2:
        raise HTTPException(400, "month must be YYYY-MM format")
    year, mo = int(parts[0]), int(parts[1])
    start = f"{year}-{mo:02d}-01"
    last_day = calendar.monthrange(year, mo)[1]
    end = f"{year}-{mo:02d}-{last_day}"

    # Actual executed transactions in this month
    txs = db.query(models.Transaction).filter(
        models.Transaction.project_id == project_id,
        models.Transaction.date >= start,
        models.Transaction.date <= end,
        models.Transaction.status == 'executed'
    ).all()

    inflow_items = []
    outflow_items = []
    for tx in txs:
        cp = db.query(models.Counterparty).filter(models.Counterparty.id == tx.counterparty_id).first() if tx.counterparty_id else None
        cat_name = tx.category or "Uncategorized"
        if tx.budget_item_id:
            cat_obj = db.query(models.BudgetCategory).filter(models.BudgetCategory.id == tx.budget_item_id).first()
            if cat_obj:
                cat_name = cat_obj.category_name
        item = {
            "date": str(tx.date) if tx.date else "",
            "category": cat_name,
            "counterparty": cp.name if cp else tx.supplier or "Unknown",
            "amount": float(tx.amount or 0),
            "status": "executed",
            "reference": tx.source_ref or tx.description or ""
        }
        if tx.direction == 'in':
            inflow_items.append(item)
        else:
            outflow_items.append(item)

    # Planned (not yet executed) for this month
    planned_txs = db.query(models.Transaction).filter(
        models.Transaction.project_id == project_id,
        models.Transaction.date >= start,
        models.Transaction.date <= end,
        models.Transaction.status == 'planned'
    ).all()

    for tx in planned_txs:
        cp = db.query(models.Counterparty).filter(models.Counterparty.id == tx.counterparty_id).first() if tx.counterparty_id else None
        cat_name = tx.category or "Uncategorized"
        if tx.budget_item_id:
            cat_obj = db.query(models.BudgetCategory).filter(models.BudgetCategory.id == tx.budget_item_id).first()
            if cat_obj:
                cat_name = cat_obj.category_name
        item = {
            "date": str(tx.date) if tx.date else "",
            "category": cat_name,
            "counterparty": cp.name if cp else tx.supplier or "Unknown",
            "amount": float(tx.amount or 0),
            "status": "planned",
            "reference": tx.source_ref or tx.description or ""
        }
        if tx.direction == 'in':
            inflow_items.append(item)
        else:
            outflow_items.append(item)

    # Expected collections (unpaid apartment balances)
    from services.forecast_service import compute_unpaid_balances
    unpaid_list = compute_unpaid_balances(db, project_id)
    for entry in unpaid_list:
        unpaid_amt = float(entry.get("unpaid", 0))
        if unpaid_amt > 0:
            apt = db.query(models.Apartment).filter(models.Apartment.id == entry["apartment_id"]).first()
            cust_name = "Unknown"
            if apt and apt.customer_id:
                cust = db.query(models.Customer).filter(models.Customer.id == apt.customer_id).first()
                if cust:
                    cust_name = cust.full_name
            elif apt and apt.customer_name:
                cust_name = apt.customer_name
            inflow_items.append({
                "date": start,
                "category": "Expected Collection",
                "counterparty": f"{cust_name} - {entry.get('apartment_name', '')}",
                "amount": unpaid_amt,
                "status": "expected",
                "reference": "Unpaid apartment balance"
            })

    return {
        "month": month,
        "project_id": project_id,
        "inflow_items": sorted(inflow_items, key=lambda x: x.get("date", "")),
        "outflow_items": sorted(outflow_items, key=lambda x: x.get("date", "")),
        "inflow_total": sum(i["amount"] for i in inflow_items),
        "outflow_total": sum(i["amount"] for i in outflow_items)
    }

@app.get("/projects/{project_id}/budget-items", response_model=List[schemas.BudgetCategory])
def read_project_budget_items(project_id: int, db: Session = Depends(get_db)):
    items = db.query(models.BudgetCategory).filter(models.BudgetCategory.project_id == project_id).all()
    return items

@app.put("/budget-categories/{category_id}", response_model=schemas.BudgetCategory)
def update_budget_category(category_id: int, update: schemas.BudgetCategoryUpdate, db: Session = Depends(get_db)):
    db_category = db.query(models.BudgetCategory).filter(models.BudgetCategory.id == category_id).first()
    if not db_category:
        raise HTTPException(status_code=404, detail="Budget category not found")
    if update.planned_amount is not None:
        db_category.planned_amount = update.planned_amount
    if update.category_name is not None:
        db_category.category_name = update.category_name
    db.commit()
    db.refresh(db_category)
    return db_category

# --- Accounts ---

@app.get("/accounts/", response_model=List[schemas.Account])
def read_accounts(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """רשימת כל החשבונות"""
    try:
        accounts = db.query(models.Account).offset(skip).limit(limit).all()
        return accounts
    except Exception as e:
        print(f"Error fetching accounts: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# --- Apartments ---

@app.get("/projects/{project_id}/apartments")
def read_apartments(project_id: int, skip: int = 0, limit: int = 50, db: Session = Depends(get_db)):
    query = db.query(models.Apartment).filter(models.Apartment.project_id == project_id)
    total = query.count()
    apartments = query.offset(skip).limit(limit).all()
    result = []
    for apt in apartments:
        total_paid = db.query(func.sum(models.CustomerPayment.amount)).filter(
            models.CustomerPayment.apartment_id == apt.id
        ).scalar() or 0
        sale_price = float(apt.sale_price) if apt.sale_price else None
        remaining = (sale_price - float(total_paid)) if sale_price is not None else None
        result.append({
            "id": apt.id,
            "project_id": apt.project_id,
            "name": apt.name,
            "floor": apt.floor,
            "apartment_number": apt.apartment_number,
            "customer_name": apt.customer_name,
            "customer_key": apt.customer_key,
            "sale_price": sale_price,
            "ownership_percent": float(apt.ownership_percent) if apt.ownership_percent else None,
            "remarks": apt.remarks,
            "total_paid": float(total_paid),
            "remaining": remaining,
        })
    return {"items": result, "total": total, "skip": skip, "limit": limit}

@app.post("/projects/{project_id}/apartments", response_model=schemas.Apartment)
def create_apartment(project_id: int, apartment: schemas.ApartmentCreate, db: Session = Depends(get_db)):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    db_apartment = models.Apartment(project_id=project_id, **apartment.dict())
    db.add(db_apartment)
    db.commit()
    db.refresh(db_apartment)
    return {**db_apartment.__dict__, "total_paid": 0, "remaining": float(db_apartment.sale_price) if db_apartment.sale_price else None}

@app.put("/apartments/{apartment_id}", response_model=schemas.Apartment)
def update_apartment(apartment_id: int, apartment: schemas.ApartmentCreate, db: Session = Depends(get_db)):
    db_apartment = db.query(models.Apartment).filter(models.Apartment.id == apartment_id).first()
    if not db_apartment:
        raise HTTPException(status_code=404, detail="Apartment not found")
    for key, value in apartment.dict().items():
        setattr(db_apartment, key, value)
    db.commit()
    db.refresh(db_apartment)
    total_paid = db.query(func.sum(models.CustomerPayment.amount)).filter(
        models.CustomerPayment.apartment_id == apartment_id
    ).scalar() or 0
    sale_price = float(db_apartment.sale_price) if db_apartment.sale_price else None
    remaining = (sale_price - float(total_paid)) if sale_price is not None else None
    return {**db_apartment.__dict__, "total_paid": float(total_paid), "remaining": remaining}

@app.delete("/apartments/{apartment_id}")
def delete_apartment(apartment_id: int, db: Session = Depends(get_db)):
    db_apartment = db.query(models.Apartment).filter(models.Apartment.id == apartment_id).first()
    if not db_apartment:
        raise HTTPException(status_code=404, detail="Apartment not found")
    db.delete(db_apartment)
    db.commit()
    return {"message": "Apartment deleted successfully"}

# --- Customer Payments ---

@app.get("/apartments/{apartment_id}/payments", response_model=List[schemas.CustomerPayment])
def read_payments(apartment_id: int, db: Session = Depends(get_db)):
    payments = db.query(models.CustomerPayment).filter(
        models.CustomerPayment.apartment_id == apartment_id
    ).order_by(models.CustomerPayment.date.desc()).all()
    return payments

@app.post("/apartments/{apartment_id}/payments", response_model=schemas.CustomerPayment)
def create_payment(apartment_id: int, payment: schemas.CustomerPaymentCreate, db: Session = Depends(get_db)):
    apartment = db.query(models.Apartment).filter(models.Apartment.id == apartment_id).first()
    if not apartment:
        raise HTTPException(status_code=404, detail="Apartment not found")
    db_payment = models.CustomerPayment(apartment_id=apartment_id, **payment.dict())
    db.add(db_payment)
    db.commit()
    db.refresh(db_payment)
    return db_payment

@app.put("/payments/{payment_id}", response_model=schemas.CustomerPayment)
def update_payment(payment_id: int, payment: schemas.CustomerPaymentCreate, db: Session = Depends(get_db)):
    db_payment = db.query(models.CustomerPayment).filter(models.CustomerPayment.id == payment_id).first()
    if not db_payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    for key, value in payment.dict().items():
        setattr(db_payment, key, value)
    db.commit()
    db.refresh(db_payment)
    return db_payment

@app.delete("/payments/{payment_id}")
def delete_payment(payment_id: int, db: Session = Depends(get_db)):
    db_payment = db.query(models.CustomerPayment).filter(models.CustomerPayment.id == payment_id).first()
    if not db_payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    db.delete(db_payment)
    db.commit()
    return {"message": "Payment deleted successfully"}

# --- Budget Plans ---

@app.get("/budget-categories/{category_id}/plans", response_model=List[schemas.BudgetPlan])
def read_budget_plans(category_id: int, db: Session = Depends(get_db)):
    plans = db.query(models.BudgetPlan).filter(
        models.BudgetPlan.budget_category_id == category_id
    ).order_by(models.BudgetPlan.planned_date).all()
    return plans

@app.post("/budget-categories/{category_id}/plans", response_model=schemas.BudgetPlan)
def create_budget_plan(category_id: int, plan: schemas.BudgetPlanCreate, db: Session = Depends(get_db)):
    category = db.query(models.BudgetCategory).filter(models.BudgetCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Budget category not found")
    db_plan = models.BudgetPlan(budget_category_id=category_id, **plan.dict())
    db.add(db_plan)
    db.commit()
    db.refresh(db_plan)
    return db_plan

@app.put("/budget-plans/{plan_id}", response_model=schemas.BudgetPlan)
def update_budget_plan(plan_id: int, plan: schemas.BudgetPlanCreate, db: Session = Depends(get_db)):
    db_plan = db.query(models.BudgetPlan).filter(models.BudgetPlan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Budget plan not found")
    for key, value in plan.dict().items():
        setattr(db_plan, key, value)
    db.commit()
    db.refresh(db_plan)
    return db_plan

@app.delete("/budget-plans/{plan_id}")
def delete_budget_plan(plan_id: int, db: Session = Depends(get_db)):
    db_plan = db.query(models.BudgetPlan).filter(models.BudgetPlan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Budget plan not found")
    db.delete(db_plan)
    db.commit()
    return {"message": "Budget plan deleted successfully"}

# --- Portfolio Summary ---

@app.get("/reports/portfolio-summary")
def get_portfolio_summary(db: Session = Depends(get_db)):
    """Aggregated portfolio summary across all active projects."""
    projects = db.query(models.Project).filter(
        models.Project.status.in_(["Active", "Completed"])
    ).all()

    project_summaries = []
    total_budget_all = 0
    total_spent_all = 0
    total_collected_all = 0
    total_revenue_all = 0

    for project in projects:
        # Budget: sum of planned amounts
        total_budget = db.query(func.sum(models.BudgetCategory.planned_amount)).filter(
            models.BudgetCategory.project_id == project.id
        ).scalar() or 0
        total_budget = float(total_budget)

        # Actual spending: sum of expense transactions (executed only)
        actual_spent = db.query(func.sum(models.Transaction.amount)).filter(
            models.Transaction.project_id == project.id,
            models.Transaction.transaction_type == 1,
            models.Transaction.type == "expense"
        ).scalar() or 0
        actual_spent = float(actual_spent)

        # Apartments: collection data
        apartments = db.query(models.Apartment).filter(
            models.Apartment.project_id == project.id
        ).all()

        total_revenue = 0
        total_collected = 0
        apartments_count = len(apartments)
        fully_paid = 0

        for apt in apartments:
            sale_price = float(apt.sale_price) if apt.sale_price else 0
            paid = db.query(func.sum(models.CustomerPayment.amount)).filter(
                models.CustomerPayment.apartment_id == apt.id
            ).scalar() or 0
            paid = float(paid)
            total_revenue += sale_price
            total_collected += paid
            if sale_price > 0 and paid >= sale_price:
                fully_paid += 1

        collection_rate = (total_collected / total_revenue * 100) if total_revenue > 0 else 0
        budget_progress = (actual_spent / total_budget * 100) if total_budget > 0 else 0

        # Budget health: count categories by status
        budget_items = db.query(models.BudgetCategory).filter(
            models.BudgetCategory.project_id == project.id
        ).all()

        categories_ok = 0
        categories_warning = 0
        categories_over = 0
        worst_category = None
        worst_overrun = 0

        for cat in budget_items:
            planned = float(cat.planned_amount) if cat.planned_amount else 0
            if planned <= 0:
                continue
            cat_actual = db.query(func.sum(models.Transaction.amount)).filter(
                models.Transaction.budget_item_id == cat.id,
                models.Transaction.transaction_type == 1
            ).scalar() or 0
            cat_actual = float(cat_actual)
            cat_progress = (cat_actual / planned * 100) if planned > 0 else 0

            if cat_progress > 100:
                categories_over += 1
                overrun = cat_actual - planned
                if overrun > worst_overrun:
                    worst_overrun = overrun
                    worst_category = {"name": cat.category_name, "progress": round(cat_progress, 1), "overrun": round(overrun, 2)}
            elif cat_progress > 90:
                categories_warning += 1
            else:
                categories_ok += 1

        total_cats = categories_ok + categories_warning + categories_over
        budget_health = round(max(0, 100 - (categories_over * 20) - (categories_warning * 5)), 0) if total_cats > 0 else 100

        # Cash flow for this project
        try:
            cash_flow = services.forecast_service.generate_cash_flow_forecast(db, project.id)
            net_cash_flow = sum(row.get("net_flow", 0) for row in cash_flow)
        except Exception:
            cash_flow = []
            net_cash_flow = 0

        project_summaries.append({
            "id": project.id,
            "name": project.name,
            "status": project.status,
            "total_budget": round(total_budget, 2),
            "actual_spent": round(actual_spent, 2),
            "budget_progress": round(budget_progress, 1),
            "total_revenue": round(total_revenue, 2),
            "total_collected": round(total_collected, 2),
            "collection_rate": round(collection_rate, 1),
            "apartments_count": apartments_count,
            "fully_paid": fully_paid,
            "net_cash_flow": round(net_cash_flow, 2),
            "budget_health": budget_health,
            "categories_ok": categories_ok,
            "categories_warning": categories_warning,
            "categories_over": categories_over,
            "worst_category": worst_category,
            "cash_flow": cash_flow,
        })

        total_budget_all += total_budget
        total_spent_all += actual_spent
        total_collected_all += total_collected
        total_revenue_all += total_revenue

    overall_collection = (total_collected_all / total_revenue_all * 100) if total_revenue_all > 0 else 0
    overall_budget_progress = (total_spent_all / total_budget_all * 100) if total_budget_all > 0 else 0

    # Feature 4: Buffer alerts
    buffer_alerts = []
    for proj_summary in project_summaries:
        setting = db.query(models.ProjectSetting).filter(
            models.ProjectSetting.project_id == proj_summary["id"]
        ).first()
        buffer_amount = float(setting.cash_buffer_amount) if setting else 200000

        proj_buffer_alerts = []
        for row in (proj_summary.get("cash_flow") or []):
            cum_balance = row.get("cumulative_balance", 0)
            if cum_balance < buffer_amount:
                shortfall = buffer_amount - cum_balance
                proj_buffer_alerts.append({
                    "month": row["date"],
                    "balance": round(cum_balance, 2),
                    "buffer": round(buffer_amount, 2),
                    "shortfall": round(shortfall, 2),
                })

        proj_summary["buffer_alerts"] = proj_buffer_alerts
        for alert in proj_buffer_alerts:
            buffer_alerts.append({
                "project_id": proj_summary["id"],
                "project_name": proj_summary["name"],
                **alert,
            })

    return {
        "projects": project_summaries,
        "totals": {
            "project_count": len(project_summaries),
            "total_budget": round(total_budget_all, 2),
            "total_spent": round(total_spent_all, 2),
            "budget_progress": round(overall_budget_progress, 1),
            "total_revenue": round(total_revenue_all, 2),
            "total_collected": round(total_collected_all, 2),
            "collection_rate": round(overall_collection, 1),
        },
        "buffer_alerts": buffer_alerts,
    }

# --- Project KPI Summary ---

@app.get("/projects/{project_id}/kpi-summary")
def get_project_kpi_summary(project_id: int, db: Session = Depends(get_db)):
    """Per-project KPI summary: collection, budget health, next month projection."""
    # Collection data from apartments
    apartments = db.query(models.Apartment).filter(
        models.Apartment.project_id == project_id
    ).all()

    total_revenue = 0
    total_collected = 0
    fully_paid = 0
    outstanding = 0

    for apt in apartments:
        sale_price = float(apt.sale_price) if apt.sale_price else 0
        paid = db.query(func.sum(models.CustomerPayment.amount)).filter(
            models.CustomerPayment.apartment_id == apt.id
        ).scalar() or 0
        paid = float(paid)
        total_revenue += sale_price
        total_collected += paid
        if sale_price > 0 and paid >= sale_price:
            fully_paid += 1
        elif sale_price > 0:
            outstanding += 1

    collection_percent = (total_collected / total_revenue * 100) if total_revenue > 0 else 0

    # Budget health
    budget_items = db.query(models.BudgetCategory).filter(
        models.BudgetCategory.project_id == project_id
    ).all()

    categories_ok = 0
    categories_warning = 0
    categories_over = 0
    worst_category = None
    worst_overrun = 0
    total_budget = 0
    total_actual_spent = 0

    for cat in budget_items:
        planned = float(cat.planned_amount) if cat.planned_amount else 0
        total_budget += planned
        if planned <= 0:
            continue
        cat_actual = db.query(func.sum(models.Transaction.amount)).filter(
            models.Transaction.budget_item_id == cat.id,
            models.Transaction.transaction_type == 1
        ).scalar() or 0
        cat_actual = float(cat_actual)
        total_actual_spent += cat_actual
        cat_progress = (cat_actual / planned * 100) if planned > 0 else 0

        if cat_progress > 100:
            categories_over += 1
            overrun = cat_actual - planned
            if overrun > worst_overrun:
                worst_overrun = overrun
                worst_category = {"name": cat.category_name, "progress": round(cat_progress, 1), "overrun": round(overrun, 2)}
        elif cat_progress > 90:
            categories_warning += 1
        else:
            categories_ok += 1

    total_cats = categories_ok + categories_warning + categories_over
    budget_health = round(max(0, 100 - (categories_over * 20) - (categories_warning * 5)), 0) if total_cats > 0 else 100

    # Next month projection from cash flow
    from datetime import datetime
    now = datetime.now()
    next_month = now.month + 1
    next_year = now.year
    if next_month > 12:
        next_month = 1
        next_year += 1
    next_month_key = f"{next_year}-{next_month:02d}"

    try:
        cash_flow = services.forecast_service.generate_cash_flow_forecast(db, project_id)
        next_month_data = next((row for row in cash_flow if row["date"] == next_month_key), None)
    except Exception:
        next_month_data = None

    next_month_income = 0
    next_month_expense = 0
    if next_month_data:
        next_month_income = next_month_data.get("actual_income", 0) + next_month_data.get("planned_income", 0)
        next_month_expense = next_month_data.get("actual_expense", 0) + next_month_data.get("planned_expense", 0)

    return {
        "collection": {
            "total_revenue": round(total_revenue, 2),
            "total_collected": round(total_collected, 2),
            "collection_percent": round(collection_percent, 1),
            "fully_paid": fully_paid,
            "outstanding": outstanding,
            "total_apartments": len(apartments),
        },
        "budget_health": {
            "score": budget_health,
            "categories_ok": categories_ok,
            "categories_warning": categories_warning,
            "categories_over": categories_over,
            "worst_category": worst_category,
            "total_budget": round(total_budget, 2),
            "total_spent": round(total_actual_spent, 2),
        },
        "next_month": {
            "month": next_month_key,
            "projected_income": round(next_month_income, 2),
            "projected_expense": round(next_month_expense, 2),
            "gap": round(next_month_income - next_month_expense, 2),
        }
    }

# --- Budget Timeline ---

@app.get("/reports/budget-timeline/{project_id}")
def get_budget_timeline(project_id: int, db: Session = Depends(get_db)):
    """Budget timeline: monthly planned vs actual spending per category."""
    from collections import defaultdict

    categories = db.query(models.BudgetCategory).filter(
        models.BudgetCategory.project_id == project_id
    ).all()

    if not categories:
        return []

    result = []

    for cat in categories:
        planned = float(cat.planned_amount) if cat.planned_amount else 0

        # Get budget plans (planned spending schedule)
        plans = db.query(models.BudgetPlan).filter(
            models.BudgetPlan.budget_category_id == cat.id
        ).order_by(models.BudgetPlan.planned_date).all()

        planned_by_month = defaultdict(float)
        for bp in plans:
            if bp.planned_date:
                month_key = bp.planned_date.strftime("%Y-%m")
                planned_by_month[month_key] += float(bp.amount) if bp.amount else 0

        # Get actual transactions for this category
        txns = db.query(models.Transaction).filter(
            models.Transaction.budget_item_id == cat.id,
            models.Transaction.transaction_type == 1
        ).all()

        actual_by_month = defaultdict(float)
        total_actual = 0
        for tx in txns:
            if tx.date:
                month_key = tx.date.strftime("%Y-%m")
                amt = float(tx.amount) if tx.amount else 0
                actual_by_month[month_key] += amt
                total_actual += amt

        # Fallback: match by category name if no budget_item_id matches
        if total_actual == 0 and cat.category_name:
            fallback_txns = db.query(models.Transaction).filter(
                models.Transaction.project_id == project_id,
                models.Transaction.transaction_type == 1,
                func.lower(models.Transaction.category) == cat.category_name.strip().lower()
            ).all()
            for tx in fallback_txns:
                if tx.date:
                    month_key = tx.date.strftime("%Y-%m")
                    amt = float(tx.amount) if tx.amount else 0
                    actual_by_month[month_key] += amt
                    total_actual += amt

        # Collect all months
        all_months = sorted(set(list(planned_by_month.keys()) + list(actual_by_month.keys())))

        monthly = []
        cumulative_planned = 0
        cumulative_actual = 0
        for month in all_months:
            p = planned_by_month.get(month, 0)
            a = actual_by_month.get(month, 0)
            cumulative_planned += p
            cumulative_actual += a
            monthly.append({
                "month": month,
                "planned": round(p, 2),
                "actual": round(a, 2),
                "cumulative_planned": round(cumulative_planned, 2),
                "cumulative_actual": round(cumulative_actual, 2),
            })

        progress = (total_actual / planned * 100) if planned > 0 else 0

        result.append({
            "id": cat.id,
            "name": cat.category_name,
            "budget": round(planned, 2),
            "total_actual": round(total_actual, 2),
            "progress": round(progress, 1),
            "variance": round(planned - total_actual, 2),
            "monthly": monthly,
            "start_month": all_months[0] if all_months else None,
            "end_month": all_months[-1] if all_months else None,
        })

    return result

# --- CSV Import ---

@app.post("/import/apartments")
async def import_apartments(file: UploadFile = File(...), db: Session = Depends(get_db)):
    from services.apartment_import_service import import_apartments_from_csv
    try:
        content = await file.read()
        result = import_apartments_from_csv(db, content)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Feature 4: Project Settings (Cash Buffer) ---

@app.get("/projects/{project_id}/settings")
def get_project_settings(project_id: int, db: Session = Depends(get_db)):
    setting = db.query(models.ProjectSetting).filter(
        models.ProjectSetting.project_id == project_id
    ).first()
    if not setting:
        return {"project_id": project_id, "cash_buffer_amount": 200000}
    return {
        "id": setting.id,
        "project_id": setting.project_id,
        "cash_buffer_amount": float(setting.cash_buffer_amount) if setting.cash_buffer_amount else 200000,
    }

@app.put("/projects/{project_id}/settings")
def update_project_settings(project_id: int, settings: schemas.ProjectSettingCreate, db: Session = Depends(get_db)):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    existing = db.query(models.ProjectSetting).filter(
        models.ProjectSetting.project_id == project_id
    ).first()
    if existing:
        existing.cash_buffer_amount = settings.cash_buffer_amount
    else:
        existing = models.ProjectSetting(
            project_id=project_id,
            cash_buffer_amount=settings.cash_buffer_amount
        )
        db.add(existing)
    db.commit()
    db.refresh(existing)
    return {
        "id": existing.id,
        "project_id": existing.project_id,
        "cash_buffer_amount": float(existing.cash_buffer_amount) if existing.cash_buffer_amount else 200000,
    }

# --- Feature 3: Suggested Category ---

def _upsert_account_category_mapping(db: Session, tx):
    """Upsert AccountCategoryMapping when transaction has both to_account_id and budget_item_id."""
    if tx.to_account_id and tx.budget_item_id:
        existing = db.query(models.AccountCategoryMapping).filter(
            models.AccountCategoryMapping.account_id == tx.to_account_id,
            models.AccountCategoryMapping.budget_category_id == tx.budget_item_id,
        ).first()
        if existing:
            existing.last_used = datetime.now()
        else:
            mapping = models.AccountCategoryMapping(
                account_id=tx.to_account_id,
                budget_category_id=tx.budget_item_id,
                last_used=datetime.now(),
            )
            db.add(mapping)
        db.commit()

@app.get("/accounts/{account_id}/suggested-category")
def get_suggested_category(account_id: int, db: Session = Depends(get_db)):
    mapping = db.query(models.AccountCategoryMapping).filter(
        models.AccountCategoryMapping.account_id == account_id
    ).order_by(models.AccountCategoryMapping.last_used.desc()).first()
    if not mapping:
        return {"budget_category_id": None}
    return {"budget_category_id": mapping.budget_category_id}

# --- Feature 1: Apartment Search ---

@app.get("/apartments/search")
def search_apartments(
    q: str = Query("", min_length=0),
    project_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    if len(q) < 2:
        return []
    query = db.query(models.Apartment).filter(
        models.Apartment.customer_name.ilike(f"%{q}%")
    )
    if project_id:
        query = query.filter(models.Apartment.project_id == project_id)
    results = query.limit(10).all()
    return [
        {
            "id": apt.id,
            "name": apt.name,
            "customer_name": apt.customer_name,
            "project_id": apt.project_id,
        }
        for apt in results
    ]

# --- Feature 5: Direct to Owner Payment ---

@app.post("/apartments/{apartment_id}/payments/direct-to-owner")
def create_direct_to_owner_payment(
    apartment_id: int,
    payment: schemas.CustomerPaymentCreate,
    db: Session = Depends(get_db),
):
    apartment = db.query(models.Apartment).filter(models.Apartment.id == apartment_id).first()
    if not apartment:
        raise HTTPException(status_code=404, detail="Apartment not found")

    # Find Direct Account and Owner Account
    direct_account = db.query(models.Account).filter(
        models.Account.name.ilike("%direct%"),
        models.Account.is_system_account == 1,
    ).first()
    owner_account = db.query(models.Account).filter(
        models.Account.name.ilike("%owner%"),
    ).first()

    if not direct_account:
        # Check if a "Direct" account exists but isn't marked as system
        maybe_direct = db.query(models.Account).filter(
            models.Account.name.ilike("%direct%")
        ).first()
        if maybe_direct:
            raise HTTPException(
                status_code=400,
                detail=f"Found account '{maybe_direct.name}' (id={maybe_direct.id}) but is_system_account is not set. Please set is_system_account=1 on this account."
            )
        raise HTTPException(status_code=400, detail="No account with 'Direct' in the name exists. Please create a system account with 'Direct' in the name.")
    if not owner_account:
        raise HTTPException(status_code=400, detail="No account with 'Owner' in the name exists. Please create an account with 'Owner' in the name.")

    customer_name = apartment.customer_name or "Unknown"

    try:
        # Create CustomerPayment record
        db_payment = models.CustomerPayment(
            apartment_id=apartment_id,
            date=payment.date,
            amount=payment.amount,
            payment_method="Direct to Owner",
            notes=payment.notes,
        )
        db.add(db_payment)
        db.flush()

        # TX1: Income to Direct Account
        tx1 = models.Transaction(
            project_id=apartment.project_id,
            date=payment.date,
            amount=payment.amount,
            to_account_id=direct_account.id,
            remarks=f"Direct to Owner - {customer_name} - IN",
            transaction_type=1,
            type="income",
            apartment_id=apartment_id,
        )
        db.add(tx1)
        db.flush()

        # TX2: Expense from Direct Account to Owner Account
        tx2 = models.Transaction(
            project_id=apartment.project_id,
            date=payment.date,
            amount=payment.amount,
            from_account_id=direct_account.id,
            to_account_id=owner_account.id,
            remarks=f"Direct to Owner - {customer_name} - OUT",
            transaction_type=1,
            type="expense",
            apartment_id=apartment_id,
        )
        db.add(tx2)
        db.flush()

        # Store linked transaction IDs
        db_payment.linked_transaction_ids = json.dumps([tx1.id, tx2.id])

        db.commit()
        db.refresh(db_payment)
        db.refresh(tx1)
        db.refresh(tx2)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create Direct to Owner payment: {str(e)}")

    return {
        "payment": {
            "id": db_payment.id,
            "apartment_id": db_payment.apartment_id,
            "date": db_payment.date.isoformat() if db_payment.date else None,
            "amount": float(db_payment.amount),
            "payment_method": db_payment.payment_method,
            "notes": db_payment.notes,
            "linked_transaction_ids": db_payment.linked_transaction_ids,
        },
        "transactions_created": 2,
    }


# --- Diagnostics ---

@app.get("/diagnostics/system-accounts")
def diagnostics_system_accounts(db: Session = Depends(get_db)):
    """Returns all accounts with system account flags, highlights Direct and Owner candidates."""
    all_accounts = db.query(models.Account).all()
    accounts_list = []
    direct_candidate = None
    owner_candidate = None

    for acc in all_accounts:
        entry = {
            "id": acc.id,
            "name": acc.name,
            "is_system_account": bool(acc.is_system_account),
        }
        if "direct" in (acc.name or "").lower():
            entry["role"] = "Direct Account candidate"
            if acc.is_system_account:
                direct_candidate = acc
        if "owner" in (acc.name or "").lower():
            entry["role"] = "Owner Account candidate"
            owner_candidate = acc
        accounts_list.append(entry)

    issues = []
    if not direct_candidate:
        # Check if there's a non-system Direct account
        maybe = next((a for a in all_accounts if "direct" in (a.name or "").lower()), None)
        if maybe:
            issues.append(f"Account '{maybe.name}' (id={maybe.id}) found but is_system_account is not set")
        else:
            issues.append("No account with 'Direct' in the name exists")
    if not owner_candidate:
        issues.append("No account with 'Owner' in the name exists")

    return {
        "accounts": accounts_list,
        "status": "ok" if not issues else "misconfigured",
        "issues": issues,
    }


# --- Admin: Bulk Budget Mapper ---

def _normalize(s):
    """Normalize string for matching: lowercase, strip, collapse whitespace."""
    if not s:
        return ""
    return " ".join(str(s).strip().lower().split())

def _match_transaction_to_category(tx, budget_categories):
    """
    Try to match a transaction to a budget category using keyword heuristics.
    Returns (category_id, category_name, match_method) or (None, None, None).
    """
    # Collect searchable text from the transaction
    tx_category = _normalize(tx.category)
    tx_description = _normalize(tx.description)
    tx_remarks = _normalize(tx.remarks)

    best_match = None
    best_score = 0

    for cat in budget_categories:
        cat_name_norm = _normalize(cat.category_name)
        if not cat_name_norm:
            continue

        score = 0
        method = None

        # 1. Exact match on legacy category field (highest confidence)
        if tx_category and tx_category == cat_name_norm:
            score = 100
            method = "exact_category"

        # 2. Category field contains the budget category name
        elif tx_category and cat_name_norm in tx_category:
            score = 80
            method = "category_contains"

        # 3. Budget category name found in description or remarks
        elif tx_description and cat_name_norm in tx_description:
            score = 60
            method = "description_contains"
        elif tx_remarks and cat_name_norm in tx_remarks:
            score = 60
            method = "remarks_contains"

        else:
            # 4. Keyword matching: all words from category name appear in any field
            cat_words = cat_name_norm.split()
            if len(cat_words) > 0:
                combined_text = f"{tx_category} {tx_description} {tx_remarks}"
                matches = sum(1 for w in cat_words if w in combined_text)
                if matches == len(cat_words):
                    score = 40
                    method = "keyword_match"

        if score > best_score:
            best_score = score
            best_match = (cat.id, cat.category_name, method)

    return best_match if best_match else (None, None, None)


def _resolve_to_account(tx, db):
    """Get the to_account display name from counterparty or account."""
    if tx.counterparty_id:
        cp = db.query(models.Counterparty).filter(models.Counterparty.id == tx.counterparty_id).first()
        if cp:
            return cp.name
    if tx.to_account_id:
        acc = db.query(models.Account).filter(models.Account.id == tx.to_account_id).first()
        if acc:
            return acc.name
    return tx.supplier or ""


@app.post("/admin/budget-mapper/{project_id}")
def bulk_budget_mapper(
    project_id: int,
    dry_run: bool = Query(True, description="If true, preview only; if false, commit changes"),
    db: Session = Depends(get_db),
):
    """
    Scan transactions with empty budget_item_id and attempt to map them
    to budget categories based on keywords in category/description/remarks.
    """
    # Fetch budget categories for this project
    budget_categories = db.query(models.BudgetCategory).filter(
        models.BudgetCategory.project_id == project_id
    ).all()

    if not budget_categories:
        raise HTTPException(status_code=404, detail="No budget categories found for this project")

    # Fetch unmapped transactions for this project
    unmapped_txs = db.query(models.Transaction).filter(
        models.Transaction.project_id == project_id,
        (models.Transaction.budget_item_id == None) | (models.Transaction.budget_item_id == 0),
    ).all()

    mappings = []       # Successfully matched
    unmatched = []      # Could not match
    updated_count = 0

    for tx in unmapped_txs:
        cat_id, cat_name, method = _match_transaction_to_category(tx, budget_categories)

        if cat_id:
            mappings.append({
                "transaction_id": tx.id,
                "date": tx.date.isoformat() if tx.date else None,
                "amount": float(tx.amount) if tx.amount else 0,
                "category_field": tx.category,
                "description": tx.description or tx.remarks,
                "mapped_to_id": cat_id,
                "mapped_to_name": cat_name,
                "match_method": method,
                "to_account": _resolve_to_account(tx, db),
                "direction": tx.direction or "out",
            })

            if not dry_run:
                tx.budget_item_id = cat_id
                updated_count += 1
        else:
            unmatched.append({
                "transaction_id": tx.id,
                "date": tx.date.isoformat() if tx.date else None,
                "amount": float(tx.amount) if tx.amount else 0,
                "category_field": tx.category,
                "description": tx.description or tx.remarks,
                "to_account": _resolve_to_account(tx, db),
                "direction": tx.direction or "out",
            })

    if not dry_run and updated_count > 0:
        db.commit()

    # Group mappings by category for summary
    category_summary = {}
    for m in mappings:
        name = m["mapped_to_name"]
        if name not in category_summary:
            category_summary[name] = {"count": 0, "total_amount": 0}
        category_summary[name]["count"] += 1
        category_summary[name]["total_amount"] += m["amount"]

    return {
        "dry_run": dry_run,
        "project_id": project_id,
        "total_unmapped": len(unmapped_txs),
        "total_matched": len(mappings),
        "total_unmatched": len(unmatched),
        "updated": updated_count,
        "category_summary": category_summary,
        "mappings": mappings,
        "unmatched": unmatched,
    }


@app.put("/admin/bulk-assign-budget")
def bulk_assign_budget(
    payload: schemas.BulkAssignBudget,
    db: Session = Depends(get_db),
):
    """Bulk-assign a budget category to multiple transactions."""
    if not payload.transaction_ids:
        raise HTTPException(status_code=400, detail="transaction_ids is required")

    category = db.query(models.BudgetCategory).filter(
        models.BudgetCategory.id == payload.budget_category_id
    ).first()
    if not category:
        raise HTTPException(status_code=404, detail="Budget category not found")

    update_fields = {models.Transaction.budget_item_id: payload.budget_category_id}
    if payload.direction and payload.direction in ('in', 'out'):
        update_fields[models.Transaction.direction] = payload.direction
        update_fields[models.Transaction.type] = 'income' if payload.direction == 'in' else 'expense'

    updated = db.query(models.Transaction).filter(
        models.Transaction.id.in_(payload.transaction_ids)
    ).update(update_fields, synchronize_session="fetch")

    db.commit()

    return {"updated": updated, "budget_category_id": payload.budget_category_id, "category_name": category.category_name}


@app.post("/admin/backfill-vat")
def backfill_vat(
    project_id: int,
    vat_rate: float = Query(0.24, description="VAT rate to apply (e.g. 0.24 for 24%)"),
    db: Session = Depends(get_db)
):
    """Backfill vat_amount = amount * vat_rate for transactions missing VAT data."""
    txs = db.query(models.Transaction).filter(
        models.Transaction.project_id == project_id,
        models.Transaction.direction == 'out',
        or_(models.Transaction.vat_amount == None, models.Transaction.vat_amount == 0)
    ).all()
    updated = 0
    for tx in txs:
        if tx.amount and float(tx.amount) > 0:
            tx.vat_amount = float(tx.amount) * vat_rate
            tx.vat_rate = vat_rate
            updated += 1
    db.commit()
    return {"updated": updated, "vat_rate_applied": vat_rate}


@app.post("/admin/backfill-withholding")
def backfill_withholding(
    project_id: int,
    withholding_rate: float = Query(0.03, description="Withholding rate (e.g. 0.03 for 3%)"),
    db: Session = Depends(get_db)
):
    """Backfill withholding_amount = amount * rate for transactions missing withholding data."""
    txs = db.query(models.Transaction).filter(
        models.Transaction.project_id == project_id,
        models.Transaction.direction == 'out',
        or_(models.Transaction.withholding_amount == None, models.Transaction.withholding_amount == 0)
    ).all()
    updated = 0
    for tx in txs:
        if tx.amount and float(tx.amount) > 0:
            tx.withholding_amount = float(tx.amount) * withholding_rate
            tx.withholding_rate = withholding_rate
            updated += 1
    db.commit()
    return {"updated": updated, "withholding_rate_applied": withholding_rate}


# --- Counterparties ---

@app.get("/counterparties/", response_model=List[schemas.Counterparty])
def get_counterparties(db: Session = Depends(get_db)):
    return db.query(models.Counterparty).all()

@app.post("/counterparties/", response_model=schemas.Counterparty)
def create_counterparty(counterparty: schemas.CounterpartyCreate, db: Session = Depends(get_db)):
    db_counterparty = models.Counterparty(**counterparty.dict())
    db.add(db_counterparty)
    db.commit()
    db.refresh(db_counterparty)
    return db_counterparty

@app.put("/counterparties/{counterparty_id}", response_model=schemas.Counterparty)
def update_counterparty(counterparty_id: int, counterparty: schemas.CounterpartyCreate, db: Session = Depends(get_db)):
    db_counterparty = db.query(models.Counterparty).filter(models.Counterparty.id == counterparty_id).first()
    if not db_counterparty:
        raise HTTPException(status_code=404, detail="Counterparty not found")
    for key, value in counterparty.dict().items():
        setattr(db_counterparty, key, value)
    db.commit()
    db.refresh(db_counterparty)
    return db_counterparty

@app.delete("/counterparties/{counterparty_id}")
def delete_counterparty(counterparty_id: int, db: Session = Depends(get_db)):
    db_counterparty = db.query(models.Counterparty).filter(models.Counterparty.id == counterparty_id).first()
    if not db_counterparty:
        raise HTTPException(status_code=404, detail="Counterparty not found")
    db.delete(db_counterparty)
    db.commit()
    return {"ok": True}


# --- Customers ---

@app.get("/customers/", response_model=List[schemas.Customer])
def get_customers(db: Session = Depends(get_db)):
    return db.query(models.Customer).all()

@app.post("/customers/", response_model=schemas.Customer)
def create_customer(customer: schemas.CustomerCreate, db: Session = Depends(get_db)):
    db_customer = models.Customer(**customer.dict())
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer

@app.put("/customers/{customer_id}", response_model=schemas.Customer)
def update_customer(customer_id: int, customer: schemas.CustomerCreate, db: Session = Depends(get_db)):
    db_customer = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    for key, value in customer.dict().items():
        setattr(db_customer, key, value)
    db.commit()
    db.refresh(db_customer)
    return db_customer

@app.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    db_customer = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    db.delete(db_customer)
    db.commit()
    return {"ok": True}


# --- Invoices ---

@app.post("/invoices/", response_model=schemas.Invoice)
def create_invoice(invoice: schemas.InvoiceCreate, db: Session = Depends(get_db)):
    db_invoice = models.Invoice(**invoice.dict())
    db.add(db_invoice)
    db.commit()
    db.refresh(db_invoice)
    # Write audit log
    log = models.AuditLog(
        entity_type="invoice",
        entity_id=db_invoice.id,
        action="create",
        diff_json=json.dumps(invoice.dict(), default=str)
    )
    db.add(log)
    db.commit()
    return db_invoice

@app.get("/invoices/")
def get_invoices(
    project_id: Optional[int] = None,
    customer_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Invoice)
    if project_id:
        query = query.filter(models.Invoice.project_id == project_id)
    if customer_id:
        query = query.filter(models.Invoice.customer_id == customer_id)
    if date_from:
        query = query.filter(models.Invoice.invoice_date >= date_from)
    if date_to:
        query = query.filter(models.Invoice.invoice_date <= date_to)
    invoices = query.all()
    # Return with customer/counterparty names
    result = []
    for inv in invoices:
        item = {
            "id": inv.id,
            "project_id": inv.project_id,
            "customer_id": inv.customer_id,
            "counterparty_id": inv.counterparty_id,
            "invoice_number": inv.invoice_number,
            "invoice_date": str(inv.invoice_date) if inv.invoice_date else None,
            "invoice_value": float(inv.invoice_value) if inv.invoice_value else 0,
            "currency": inv.currency,
            "remarks": inv.remarks,
            "created_at": str(inv.created_at) if inv.created_at else None,
        }
        # Add linked transaction total
        linked_txs = db.query(models.Transaction).filter(models.Transaction.invoice_id == inv.id).all()
        item["transactions_value"] = float(sum(tx.amount or 0 for tx in linked_txs))
        item["balance"] = item["invoice_value"] - item["transactions_value"]
        result.append(item)
    return result

@app.put("/invoices/{invoice_id}", response_model=schemas.Invoice)
def update_invoice(invoice_id: int, invoice: schemas.InvoiceCreate, db: Session = Depends(get_db)):
    db_invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    old_data = {c.name: str(getattr(db_invoice, c.name)) for c in models.Invoice.__table__.columns}
    for key, value in invoice.dict().items():
        setattr(db_invoice, key, value)
    db.commit()
    db.refresh(db_invoice)
    log = models.AuditLog(
        entity_type="invoice",
        entity_id=invoice_id,
        action="update",
        diff_json=json.dumps({"old": old_data, "new": invoice.dict()}, default=str)
    )
    db.add(log)
    db.commit()
    return db_invoice

@app.delete("/invoices/{invoice_id}")
def delete_invoice(invoice_id: int, db: Session = Depends(get_db)):
    db_invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    log = models.AuditLog(
        entity_type="invoice",
        entity_id=invoice_id,
        action="delete",
        diff_json=json.dumps({"invoice_number": db_invoice.invoice_number}, default=str)
    )
    db.add(log)
    db.delete(db_invoice)
    db.commit()
    return {"ok": True}

@app.post("/invoices/import")
def import_invoices(file: UploadFile = File(...), project_id: int = Form(...), db: Session = Depends(get_db)):
    """Bulk import invoices from CSV file."""
    import csv, io
    content = file.file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    imported = 0
    errors = []
    for i, row in enumerate(reader, 1):
        try:
            raw_date = row.get("invoice_date", "").strip() or None
            parsed_date = None
            if raw_date:
                from datetime import date as _date
                parsed_date = _date.fromisoformat(raw_date)
            inv = models.Invoice(
                project_id=project_id,
                invoice_number=row.get("invoice_number", "").strip(),
                invoice_date=parsed_date,
                invoice_value=Decimal(row.get("invoice_value", "0").strip().replace(",", "") or "0"),
                currency=row.get("currency", "EUR").strip() or "EUR",
            )
            # Link customer by name if provided
            cust_name = row.get("customer_name", "").strip()
            if cust_name:
                cust = db.query(models.Customer).filter(
                    func.lower(models.Customer.full_name) == cust_name.lower()
                ).first()
                if cust:
                    inv.customer_id = cust.id
            # Link counterparty by name if provided
            cp_name = row.get("counterparty_name", "").strip()
            if cp_name:
                cp = db.query(models.Counterparty).filter(
                    func.lower(models.Counterparty.name) == cp_name.lower()
                ).first()
                if cp:
                    inv.counterparty_id = cp.id
            db.add(inv)
            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")
    db.commit()
    return {"imported": imported, "errors": errors}

@app.post("/transactions/{transaction_id}/link-invoice")
def link_invoice_to_transaction(transaction_id: int, invoice_id: int, db: Session = Depends(get_db)):
    tx = db.query(models.Transaction).filter(models.Transaction.id == transaction_id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    tx.invoice_id = invoice_id
    db.commit()
    return {"ok": True, "transaction_id": transaction_id, "invoice_id": invoice_id}


# --- Reports ---

@app.get("/reports/invoices")
def get_invoice_report(
    project_id: Optional[int] = None,
    customer_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D
    date_from = _clean_date(date_from)
    date_to = _clean_date(date_to)
    query = db.query(models.Invoice)
    filters_applied = {}
    if project_id:
        query = query.filter(models.Invoice.project_id == project_id)
        filters_applied["project_id"] = project_id
    if customer_id:
        query = query.filter(models.Invoice.customer_id == customer_id)
        filters_applied["customer_id"] = customer_id
    if date_from:
        query = query.filter(models.Invoice.invoice_date >= date_from)
        filters_applied["date_from"] = date_from
    if date_to:
        query = query.filter(models.Invoice.invoice_date <= date_to)
        filters_applied["date_to"] = date_to

    invoices = query.all()

    # Group by customer
    from collections import defaultdict
    by_customer = defaultdict(lambda: {"invoice_value": D('0'), "transactions_value": D('0')})

    for inv in invoices:
        customer_name = "Unknown"
        if inv.customer_id:
            cust = db.query(models.Customer).filter(models.Customer.id == inv.customer_id).first()
            if cust:
                customer_name = cust.full_name

        inv_value = D(str(inv.invoice_value)) if inv.invoice_value else D('0')
        by_customer[customer_name]["invoice_value"] += inv_value

        linked_txs = db.query(models.Transaction).filter(models.Transaction.invoice_id == inv.id).all()
        tx_value = sum(D(str(tx.amount or 0)) for tx in linked_txs)
        by_customer[customer_name]["transactions_value"] += tx_value

    rows = []
    total_inv = D('0')
    total_tx = D('0')
    for customer, vals in by_customer.items():
        balance = vals["invoice_value"] - vals["transactions_value"]
        rows.append({
            "customer": customer,
            "invoice_value": float(vals["invoice_value"]),
            "transactions_value": float(vals["transactions_value"]),
            "balance": float(balance)
        })
        total_inv += vals["invoice_value"]
        total_tx += vals["transactions_value"]

    result = {
        "rows": rows,
        "totals": {
            "invoice_value": float(total_inv),
            "transactions_value": float(total_tx),
            "balance": float(total_inv - total_tx)
        },
        "drilldown_supported": True,
        "filters_applied": filters_applied
    }
    if format == "xlsx":
        buf = export_to_excel("Invoices Report", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=Invoices Report.xlsx"})
    return result


# --- Report 1: P&L ---

@app.get("/reports/pnl")
def get_pnl_report(
    project_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D
    from collections import OrderedDict

    date_from = _clean_date(date_from)
    date_to = _clean_date(date_to)

    query = db.query(models.Transaction)
    filters_applied = {}
    if project_id:
        query = query.filter(models.Transaction.project_id == project_id)
        filters_applied["project_id"] = project_id
    if date_from:
        query = query.filter(models.Transaction.date >= date_from)
        filters_applied["date_from"] = date_from
    if date_to:
        query = query.filter(models.Transaction.date <= date_to)
        filters_applied["date_to"] = date_to
    if status:
        query = query.filter(models.Transaction.status == status)
    else:
        query = query.filter(models.Transaction.status == 'executed')

    transactions = query.all()

    # Resolve category: budget_item > legacy category > "Uncategorized"
    def resolve_category(tx):
        if tx.budget_item_id:
            cat = db.query(models.BudgetCategory).filter(
                models.BudgetCategory.id == tx.budget_item_id
            ).first()
            if cat:
                return cat.category_name
        if tx.category and tx.category.strip():
            return tx.category.strip()
        return "Uncategorized"

    def resolve_counterparty(tx):
        if tx.counterparty_id:
            cp = db.query(models.Counterparty).filter(
                models.Counterparty.id == tx.counterparty_id
            ).first()
            if cp:
                return cp.name
        return tx.supplier or "Unknown"

    # Separate income vs expense, group by category then counterparty
    sections = {"income": OrderedDict(), "expense": OrderedDict()}

    for tx in transactions:
        section = "income" if tx.direction == 'in' else "expense"
        cat_name = resolve_category(tx)
        cp_name = resolve_counterparty(tx)
        amount = D(str(tx.amount or 0))
        vat = D(str(tx.vat_amount or 0))
        withholding = D(str(tx.withholding_amount or 0))

        if cat_name not in sections[section]:
            sections[section][cat_name] = OrderedDict()
        if cp_name not in sections[section][cat_name]:
            sections[section][cat_name][cp_name] = {
                "trans_value": D('0'), "vat_value": D('0'), "withholding_value": D('0')
            }
        sections[section][cat_name][cp_name]["trans_value"] += amount
        sections[section][cat_name][cp_name]["vat_value"] += vat
        sections[section][cat_name][cp_name]["withholding_value"] += withholding

    # Build rows with row_type
    rows = []
    grand = {"income": D('0'), "expense": D('0')}
    grand_vat = {"income": D('0'), "expense": D('0')}
    grand_wh = {"income": D('0'), "expense": D('0')}

    for section in ["income", "expense"]:
        rows.append({"row_type": "section_header", "section": section,
                      "category": section.capitalize(), "counterparty": "",
                      "trans_value": 0, "vat_value": 0, "value_no_vat": 0,
                      "withholding_value": 0, "value_no_vat_no_withholding": 0})

        for cat_name, counterparties in sections[section].items():
            cat_total = D('0'); cat_vat = D('0'); cat_wh = D('0')
            for cp_name, vals in counterparties.items():
                tv = vals["trans_value"]
                vv = vals["vat_value"]
                wv = vals["withholding_value"]
                vn = tv - vv
                vn_wh = vn - wv
                rows.append({
                    "row_type": "detail", "section": section,
                    "category": cat_name, "counterparty": cp_name,
                    "trans_value": float(tv), "vat_value": float(vv),
                    "value_no_vat": float(vn), "withholding_value": float(wv),
                    "value_no_vat_no_withholding": float(vn_wh)
                })
                cat_total += tv; cat_vat += vv; cat_wh += wv

            vn = cat_total - cat_vat
            rows.append({
                "row_type": "subtotal", "section": section,
                "category": f"Total {cat_name}", "counterparty": "",
                "trans_value": float(cat_total), "vat_value": float(cat_vat),
                "value_no_vat": float(vn), "withholding_value": float(cat_wh),
                "value_no_vat_no_withholding": float(vn - cat_wh)
            })
            grand[section] += cat_total
            grand_vat[section] += cat_vat
            grand_wh[section] += cat_wh

        vn = grand[section] - grand_vat[section]
        rows.append({
            "row_type": "total", "section": section,
            "category": f"Total {section.capitalize()}", "counterparty": "",
            "trans_value": float(grand[section]), "vat_value": float(grand_vat[section]),
            "value_no_vat": float(vn), "withholding_value": float(grand_wh[section]),
            "value_no_vat_no_withholding": float(vn - grand_wh[section])
        })

    # Net Profit/Loss
    net = grand["income"] - grand["expense"]
    net_vat = grand_vat["income"] - grand_vat["expense"]
    net_wh = grand_wh["income"] - grand_wh["expense"]
    net_vn = net - net_vat
    rows.append({
        "row_type": "grand_total", "section": "net",
        "category": "Net Profit / Loss", "counterparty": "",
        "trans_value": float(net), "vat_value": float(net_vat),
        "value_no_vat": float(net_vn), "withholding_value": float(net_wh),
        "value_no_vat_no_withholding": float(net_vn - net_wh)
    })

    result = {
        "rows": rows,
        "totals": {
            "income_total": float(grand["income"]),
            "expense_total": float(grand["expense"]),
            "net_profit": float(net),
            "trans_value": float(net), "vat_value": float(net_vat),
            "value_no_vat": float(net_vn),
            "withholding_value": float(net_wh),
            "value_no_vat_no_withholding": float(net_vn - net_wh)
        },
        "drilldown_supported": True,
        "filters_applied": filters_applied
    }
    if format == "xlsx":
        buf = export_to_excel("PnL Report", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=PnL Report.xlsx"})
    return result


# --- Report 2: Plan vs Actual (enhanced) ---

@app.get("/reports/plan-vs-actual")
def get_plan_vs_actual_report(
    project_id: Optional[int] = None,
    status: Optional[str] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D
    filters_applied = {}
    if project_id:
        filters_applied["project_id"] = project_id

    categories_query = db.query(models.BudgetCategory)
    if project_id:
        categories_query = categories_query.filter(models.BudgetCategory.project_id == project_id)
    categories = categories_query.all()

    rows = []
    total_plan1 = D('0')
    total_plan2 = D('0')
    total_actual = D('0')
    total_vat = D('0')
    total_withholding = D('0')

    for cat in categories:
        plan1 = D(str(cat.planned_amount or 0))
        plan2_raw = getattr(cat, 'planned_amount_v2', None)
        plan2 = D(str(plan2_raw)) if plan2_raw is not None else plan1  # Default to plan1 if no plan2

        tx_query = db.query(models.Transaction).filter(
            models.Transaction.budget_item_id == cat.id,
            models.Transaction.direction == 'out'
        )
        if status:
            tx_query = tx_query.filter(models.Transaction.status == status)
        else:
            tx_query = tx_query.filter(models.Transaction.status == 'executed')

        txs = tx_query.all()
        actual = sum(D(str(tx.amount or 0)) for tx in txs)
        vat = sum(D(str(tx.vat_amount or 0)) for tx in txs)
        withholding = sum(D(str(tx.withholding_amount or 0)) for tx in txs)

        rows.append({
            "category": cat.category_name,
            "plan1": float(plan1),
            "plan2": float(plan2),
            "plan1_plan2_diff": float(plan1 - plan2),
            "actual": float(actual),
            "plan2_actual_diff": float(plan2 - actual),
            "vat_amount": float(vat),
            "withholding_amount": float(withholding)
        })
        total_plan1 += plan1
        total_plan2 += plan2
        total_actual += actual
        total_vat += vat
        total_withholding += withholding

    result = {
        "rows": rows,
        "totals": {
            "plan1": float(total_plan1),
            "plan2": float(total_plan2),
            "plan1_plan2_diff": float(total_plan1 - total_plan2),
            "actual": float(total_actual),
            "plan2_actual_diff": float(total_plan2 - total_actual),
            "vat_amount": float(total_vat),
            "withholding_amount": float(total_withholding)
        },
        "drilldown_supported": True,
        "filters_applied": filters_applied
    }
    if format == "xlsx":
        buf = export_to_excel("Plan vs Actual", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=Plan vs Actual.xlsx"})
    return result


@app.patch("/budget-categories/{category_id}/plan2")
def update_plan2(category_id: int, amount: float = Query(...), db: Session = Depends(get_db)):
    cat = db.query(models.BudgetCategory).filter(models.BudgetCategory.id == category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    cat.planned_amount_v2 = amount
    db.commit()
    db.refresh(cat)
    return {"id": cat.id, "category_name": cat.category_name, "planned_amount_v2": float(cat.planned_amount_v2) if cat.planned_amount_v2 else None}


# --- Report 4: Customer Transactions ---

@app.get("/reports/customer-transactions")
def get_customer_transactions_report(
    project_id: Optional[int] = None,
    customer_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D

    date_from = _clean_date(date_from)
    date_to = _clean_date(date_to)
    filters_applied = {}

    # Primary: CustomerPayments via apartments (same source as Customer Balance)
    apt_query = db.query(models.Apartment).filter(models.Apartment.customer_id != None)
    if project_id:
        apt_query = apt_query.filter(models.Apartment.project_id == project_id)
        filters_applied["project_id"] = project_id
    if customer_id:
        apt_query = apt_query.filter(models.Apartment.customer_id == customer_id)
        filters_applied["customer_id"] = customer_id
    apartments = apt_query.all()

    rows = []
    total_amount = D('0')
    seen_amounts = set()  # For dedup: (customer, date, amount)

    for apt in apartments:
        cust = db.query(models.Customer).filter(models.Customer.id == apt.customer_id).first()
        proj = db.query(models.Project).filter(models.Project.id == apt.project_id).first()

        pay_query = db.query(models.CustomerPayment).filter(
            models.CustomerPayment.apartment_id == apt.id
        )
        if date_from:
            pay_query = pay_query.filter(models.CustomerPayment.date >= date_from)
        if date_to:
            pay_query = pay_query.filter(models.CustomerPayment.date <= date_to)

        payments = pay_query.all()
        for p in payments:
            amount = D(str(p.amount or 0))
            total_amount += amount
            cust_name = cust.full_name if cust else apt.customer_name or "Unknown"
            date_str = str(p.date) if p.date else ""
            seen_amounts.add((cust_name, date_str, float(amount)))
            rows.append({
                "customer": cust_name,
                "project": proj.name if proj else "Unknown",
                "apartment": apt.apartment_number or apt.unit_number or str(apt.id),
                "date": date_str,
                "amount": float(amount),
                "description": p.notes or p.payment_method or "",
                "source_ref": ""
            })

    # Secondary: Transactions with customer_id_fk (for future-linked data)
    tx_query = db.query(models.Transaction).filter(
        models.Transaction.customer_id_fk != None,
        models.Transaction.direction == 'in',
        models.Transaction.status == 'executed'
    )
    if project_id:
        tx_query = tx_query.filter(models.Transaction.project_id == project_id)
    if customer_id:
        tx_query = tx_query.filter(models.Transaction.customer_id_fk == customer_id)
    if date_from:
        tx_query = tx_query.filter(models.Transaction.date >= date_from)
    if date_to:
        tx_query = tx_query.filter(models.Transaction.date <= date_to)

    for tx in tx_query.all():
        cust = db.query(models.Customer).filter(models.Customer.id == tx.customer_id_fk).first()
        proj = db.query(models.Project).filter(models.Project.id == tx.project_id).first()
        apt = db.query(models.Apartment).filter(models.Apartment.id == tx.apartment_id).first() if tx.apartment_id else None
        amount = D(str(tx.amount or 0))
        cust_name = cust.full_name if cust else "Unknown"
        date_str = str(tx.date) if tx.date else ""

        # Skip duplicates
        dedup_key = (cust_name, date_str, float(amount))
        if dedup_key in seen_amounts:
            continue
        seen_amounts.add(dedup_key)

        total_amount += amount
        rows.append({
            "customer": cust_name,
            "project": proj.name if proj else "Unknown",
            "apartment": apt.apartment_number if apt else "",
            "date": date_str,
            "amount": float(amount),
            "description": tx.description or "",
            "source_ref": tx.source_ref or ""
        })

    rows.sort(key=lambda r: r.get("date", ""))

    result = {
        "rows": rows,
        "totals": {"amount": float(total_amount)},
        "drilldown_supported": False,
        "filters_applied": filters_applied
    }
    if format == "xlsx":
        buf = export_to_excel("Customer Transactions", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=Customer Transactions.xlsx"})
    return result


# --- Report 5: Customer Balance ---

@app.get("/reports/customer-balance")
def get_customer_balance_report(
    project_id: Optional[int] = None,
    customer_id: Optional[int] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D

    apt_query = db.query(models.Apartment).filter(models.Apartment.customer_id != None)
    filters_applied = {}
    if project_id:
        apt_query = apt_query.filter(models.Apartment.project_id == project_id)
        filters_applied["project_id"] = project_id
    if customer_id:
        apt_query = apt_query.filter(models.Apartment.customer_id == customer_id)
        filters_applied["customer_id"] = customer_id

    apartments = apt_query.all()
    rows = []
    total_price = D('0')
    total_received = D('0')

    for apt in apartments:
        cust = db.query(models.Customer).filter(models.Customer.id == apt.customer_id).first() if apt.customer_id else None
        price = D(str(apt.sale_price or 0))

        # Sum all customer payments for this apartment
        payments = db.query(models.CustomerPayment).filter(
            models.CustomerPayment.apartment_id == apt.id
        ).all()
        received = sum(D(str(p.amount or 0)) for p in payments)

        remaining = price - received
        pct_paid = float(received / price * 100) if price > 0 else 0

        rows.append({
            "customer": cust.full_name if cust else apt.customer_name or "Unknown",
            "apartment": apt.apartment_number or apt.unit_number or str(apt.id),
            "sale_price": float(price),
            "received": float(received),
            "remaining": float(remaining),
            "pct_paid": round(pct_paid, 1)
        })
        total_price += price
        total_received += received

    result = {
        "rows": rows,
        "totals": {
            "sale_price": float(total_price),
            "received": float(total_received),
            "remaining": float(total_price - total_received)
        },
        "drilldown_supported": True,
        "filters_applied": filters_applied
    }
    if format == "xlsx":
        buf = export_to_excel("Customer Balance", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=Customer Balance.xlsx"})
    return result


# --- Report 8: Payments by Project ---

@app.get("/reports/payments-by-project")
def get_payments_by_project_report(
    project_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D
    from collections import defaultdict
    date_from = _clean_date(date_from)
    date_to = _clean_date(date_to)

    query = db.query(models.Transaction).filter(models.Transaction.direction == 'out')
    filters_applied = {}
    if project_id:
        query = query.filter(models.Transaction.project_id == project_id)
        filters_applied["project_id"] = project_id
    if date_from:
        query = query.filter(models.Transaction.date >= date_from)
        filters_applied["date_from"] = date_from
    if date_to:
        query = query.filter(models.Transaction.date <= date_to)
        filters_applied["date_to"] = date_to
    if status:
        query = query.filter(models.Transaction.status == status)
        filters_applied["status"] = status

    transactions = query.all()

    rows = []
    total_amount = D('0')

    for tx in transactions:
        proj = db.query(models.Project).filter(models.Project.id == tx.project_id).first() if tx.project_id else None
        cp = db.query(models.Counterparty).filter(models.Counterparty.id == tx.counterparty_id).first() if tx.counterparty_id else None
        tx_date = tx.date
        month = str(tx_date)[:7] if tx_date else ""
        amount = D(str(tx.amount or 0))
        total_amount += amount
        rows.append({
            "project": proj.name if proj else "Unknown",
            "counterparty": cp.name if cp else (tx.supplier or "Unknown"),
            "month": month,
            "amount": float(amount),
            "description": tx.description or ""
        })

    result = {
        "rows": rows,
        "totals": {"amount": float(total_amount)},
        "drilldown_supported": False,
        "filters_applied": filters_applied
    }
    if format == "xlsx":
        buf = export_to_excel("Payments by Project", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=Payments by Project.xlsx"})
    return result


# --- Report 9: VAT ---

@app.get("/reports/vat")
def get_vat_report(
    project_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D
    date_from = _clean_date(date_from)
    date_to = _clean_date(date_to)

    query = db.query(models.Transaction).filter(models.Transaction.vat_amount > 0)
    filters_applied = {}
    if project_id:
        query = query.filter(models.Transaction.project_id == project_id)
        filters_applied["project_id"] = project_id
    if date_from:
        query = query.filter(models.Transaction.date >= date_from)
        filters_applied["date_from"] = date_from
    if date_to:
        query = query.filter(models.Transaction.date <= date_to)
        filters_applied["date_to"] = date_to

    transactions = query.all()
    total_amount = D('0')
    total_vat = D('0')
    rows = []

    for tx in transactions:
        cp = db.query(models.Counterparty).filter(models.Counterparty.id == tx.counterparty_id).first() if tx.counterparty_id else None
        amount = D(str(tx.amount or 0))
        vat = D(str(tx.vat_amount or 0))
        total_amount += amount
        total_vat += vat
        rows.append({
            "counterparty": cp.name if cp else (tx.supplier or "Unknown"),
            "description": tx.description or "",
            "date": str(tx.date) if tx.date else "",
            "amount": float(amount),
            "vat_amount": float(vat)
        })

    result = {
        "rows": rows,
        "totals": {"amount": float(total_amount), "vat_amount": float(total_vat)},
        "drilldown_supported": False,
        "filters_applied": filters_applied
    }
    if not rows:
        result["message"] = "No transactions with VAT data found. Ensure VAT amounts are populated during transaction import or entry."
    if format == "xlsx":
        buf = export_to_excel("VAT Report", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=VAT Report.xlsx"})
    return result


# --- Report 10: Withholding Tax ---

@app.get("/reports/withholding")
def get_withholding_report(
    project_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    format: str = Query(None),
    db: Session = Depends(get_db)
):
    from decimal import Decimal as D
    from collections import defaultdict
    date_from = _clean_date(date_from)
    date_to = _clean_date(date_to)

    query = db.query(models.Transaction).filter(models.Transaction.withholding_amount > 0)
    filters_applied = {}
    if project_id:
        query = query.filter(models.Transaction.project_id == project_id)
        filters_applied["project_id"] = project_id
    if date_from:
        query = query.filter(models.Transaction.date >= date_from)
        filters_applied["date_from"] = date_from
    if date_to:
        query = query.filter(models.Transaction.date <= date_to)
        filters_applied["date_to"] = date_to

    transactions = query.all()

    # Group by counterparty
    by_cp = defaultdict(lambda: {"amount": D('0'), "withholding": D('0'), "rows": []})

    for tx in transactions:
        cp = db.query(models.Counterparty).filter(models.Counterparty.id == tx.counterparty_id).first() if tx.counterparty_id else None
        cp_name = cp.name if cp else (tx.supplier or "Unknown")
        amount = D(str(tx.amount or 0))
        withholding = D(str(tx.withholding_amount or 0))
        by_cp[cp_name]["amount"] += amount
        by_cp[cp_name]["withholding"] += withholding
        by_cp[cp_name]["rows"].append({
            "counterparty": cp_name,
            "description": tx.description or "",
            "date": str(tx.date) if tx.date else "",
            "amount": float(amount),
            "withholding_amount": float(withholding)
        })

    rows = []
    total_amount = D('0')
    total_withholding = D('0')
    for cp_name, data in by_cp.items():
        rows.extend(data["rows"])
        total_amount += data["amount"]
        total_withholding += data["withholding"]

    result = {
        "rows": rows,
        "totals": {"amount": float(total_amount), "withholding_amount": float(total_withholding)},
        "drilldown_supported": False,
        "filters_applied": filters_applied
    }
    if not rows:
        result["message"] = "No transactions with withholding data found. Ensure withholding amounts are populated during transaction import or entry."
    if format == "xlsx":
        buf = export_to_excel("Withholding Tax", result["rows"], result.get("totals"), filters_applied)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=Withholding Tax.xlsx"})
    return result


@app.post("/imports/transactions")
async def import_transactions(
    file: UploadFile = File(...),
    project_id: int = Form(...),
    db: Session = Depends(get_db)
):
    """Import transactions from CSV or Excel file."""
    # Read file content
    content = await file.read()

    # Parse CSV or Excel
    if file.filename.endswith('.csv'):
        import csv, io
        reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
        rows = list(reader)
    else:
        # Excel
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, 1):
        try:
            source_ref = row.get('source_ref') or row.get('Source Ref')

            # Duplicate guard
            if source_ref:
                existing = db.query(models.Transaction).filter(
                    models.Transaction.source_ref == source_ref
                ).first()
                if existing:
                    skipped += 1
                    continue

            # Parse row
            date_str = row.get('date') or row.get('Date')
            # Parse date string to datetime for SQLite compatibility
            parsed_date = None
            if date_str:
                if isinstance(date_str, (datetime,)):
                    parsed_date = date_str
                else:
                    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
                        try:
                            parsed_date = datetime.strptime(str(date_str), fmt)
                            break
                        except ValueError:
                            continue
                    if parsed_date is None:
                        parsed_date = datetime.fromisoformat(str(date_str))
            amount_str = str(row.get('amount') or row.get('Amount') or '0')
            direction = (row.get('direction') or row.get('Direction') or 'expense').lower()
            description = row.get('description') or row.get('Description') or ''
            status = (row.get('status') or row.get('Status') or 'Executed')

            # Find category
            category_name = row.get('category') or row.get('Category')
            category = None
            if category_name:
                category = db.query(models.BudgetCategory).filter(
                    models.BudgetCategory.project_id == project_id,
                    models.BudgetCategory.category_name == category_name
                ).first()

            tx = models.Transaction(
                project_id=project_id,
                date=parsed_date,
                amount=Decimal(amount_str.replace(',', '')),
                direction=direction,
                type=direction,
                description=description,
                status=status,
                source_ref=source_ref,
                budget_item_id=category.id if category else None,
            )
            db.add(tx)
            db.flush()
            imported += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": i, "error": str(e)})

    if imported > 0:
        db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)