from io import BytesIO
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def export_to_excel(report_name: str, rows: list, totals: dict = None, filters: dict = None) -> BytesIO:
    """Export report data to Excel format, returns BytesIO buffer."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_name[:31]  # Excel sheet name max 31 chars

    # Metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws["A1"] = "Report"
    meta_ws["B1"] = report_name
    meta_ws["A2"] = "Generated"
    meta_ws["B2"] = datetime.now().isoformat()
    if filters:
        row = 3
        for k, v in filters.items():
            meta_ws.cell(row=row, column=1, value=k)
            meta_ws.cell(row=row, column=2, value=str(v) if v is not None else "")
            row += 1

    # Data: write headers from first row keys
    if rows:
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        row_idx = 1
        for row_idx, row_data in enumerate(rows, 2):
            for col, key in enumerate(headers, 1):
                val = row_data.get(key)
                if isinstance(val, Decimal):
                    val = float(val)
                ws.cell(row=row_idx, column=col, value=val)

        # Totals row if provided
        if totals:
            totals_row = row_idx + 1
            ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
            for col, key in enumerate(headers, 1):
                if key in totals:
                    val = totals[key]
                    if isinstance(val, Decimal):
                        val = float(val)
                    ws.cell(row=totals_row, column=col, value=val)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_pnl_reference_format(rows: list, totals: dict = None, filters: dict = None) -> BytesIO:
    """Export P&L in business-friendly Excel reference format.

    Layout: Category | Counterparty | Amount | VAT | Net (excl VAT) | Withholding | Net (excl VAT & WH)
    Sections styled with colored fills. No technical columns (row_type, section).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Report"

    # Styles
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    grand_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Column headers
    headers = ["Category", "Counterparty", "Amount", "VAT", "Net (excl. VAT)", "Withholding", "Net (excl. VAT & WH)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[c].width = 18

    eur_fmt = '#,##0.00'

    row_num = 2
    for row_data in rows:
        rt = row_data.get("row_type", "")
        cat = row_data.get("category", "")
        cp = row_data.get("counterparty", "")
        tv = row_data.get("trans_value", 0)
        vv = row_data.get("vat_value", 0)
        vnv = row_data.get("value_no_vat", 0)
        wv = row_data.get("withholding_value", 0)
        vnvw = row_data.get("value_no_vat_no_withholding", 0)

        if rt == "section_header":
            cell = ws.cell(row=row_num, column=1, value=cat)
            cell.font = bold
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).fill = section_fill
        elif rt == "detail":
            ws.cell(row=row_num, column=1, value=cat)
            ws.cell(row=row_num, column=2, value=cp)
            for c, v in [(3, tv), (4, vv), (5, vnv), (6, wv), (7, vnvw)]:
                cell = ws.cell(row=row_num, column=c, value=v)
                cell.number_format = eur_fmt
        elif rt == "subtotal":
            ws.cell(row=row_num, column=1, value=cat).font = bold
            for c, v in [(3, tv), (4, vv), (5, vnv), (6, wv), (7, vnvw)]:
                cell = ws.cell(row=row_num, column=c, value=v)
                cell.number_format = eur_fmt
                cell.font = bold
                cell.fill = subtotal_fill
        elif rt == "total":
            ws.cell(row=row_num, column=1, value=cat).font = bold
            for c, v in [(3, tv), (4, vv), (5, vnv), (6, wv), (7, vnvw)]:
                cell = ws.cell(row=row_num, column=c, value=v)
                cell.number_format = eur_fmt
                cell.font = bold
                cell.fill = total_fill
        elif rt == "grand_total":
            ws.cell(row=row_num, column=1, value=cat).font = bold
            for c, v in [(3, tv), (4, vv), (5, vnv), (6, wv), (7, vnvw)]:
                cell = ws.cell(row=row_num, column=c, value=v)
                cell.number_format = eur_fmt
                cell.font = bold
                cell.fill = grand_fill

        row_num += 1

    # Filters metadata at bottom
    if filters:
        row_num += 1
        ws.cell(row=row_num, column=1, value="Filters Applied:").font = bold
        for k, v in filters.items():
            row_num += 1
            ws.cell(row=row_num, column=1, value=k)
            ws.cell(row=row_num, column=2, value=str(v))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
