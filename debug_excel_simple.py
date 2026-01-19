from openpyxl import load_workbook
import shutil
import os

src = "c:/Users/iftahe/Greece Project/Progreece21 - Sheet2.csv"
dst = "c:/Users/iftahe/Greece Project/temp_debug.xlsx"

print(f"Copying {src} to {dst}...")
shutil.copyfile(src, dst)
print("Copy done.")

try:
    print("Loading workbook...")
    wb = load_workbook(dst, read_only=True)
    ws = wb.active
    print("Workbook loaded.")
    
    print("Reading header row...")
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        print(row)
        break
except Exception as e:
    print(f"Error: {e}")

# Cleanup
if os.path.exists(dst):
    try:
        os.remove(dst)
    except:
        pass
